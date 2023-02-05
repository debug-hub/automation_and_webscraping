from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl import *
import re
import pandas as pd




workbook = openpyxl.Workbook()

ws1 = workbook.create_sheet('top rated movies')
ws1 = workbook.active
ws1.title='top rated movies'

print(workbook.sheetnames)


try:
    source = requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()


    soup = BeautifulSoup(source.text,'html.parser')

    movies = soup.find('tbody', class_= 'lister-list').find_all('tr')

    ws1.append(['Movie Rank','Movie Name','Year of release','IMDB rating'])

    for movie in movies:

        name = movie.find('td',class_='titleColumn').a.text

        rank = movie.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]

        year = movie.find('td',class_='titleColumn').span.text.strip('()')

        rating = movie.find('td',class_='ratingColumn imdbRating').strong.text

        print( rank,name,year,rating)
        ws1.append([rank,name,year,rating])

    

    workbook.save(r'D:\IMDB movie rating.xlsx')

    print('DATA SAVED')

except Exception as e:
    print(e)
    


workbook2 = openpyxl.Workbook()
ws2 = workbook2.create_sheet('top rated shows')

ws2 = workbook2.active
ws2.title='top rated shows'



try:
    source = requests.get('https://www.imdb.com/chart/toptv/')
    source.raise_for_status()


    soup = BeautifulSoup(source.text,'html.parser')

    movies = soup.find('tbody', class_= 'lister-list').find_all('tr')

    ws2.append(['Show Rank','Show Name','Year of release','IMDB rating'])

    for movie in movies:

        name = movie.find('td',class_='titleColumn').a.text

        rank = movie.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]

        year = movie.find('td',class_='titleColumn').span.text.strip('()')

        rating = movie.find('td',class_='ratingColumn imdbRating').strong.text

        print( rank,name,year,rating)
        ws2.append([rank,name,year,rating])

        

    workbook2.save(r'D:\IMDB show rating.xlsx')

    print('DATA SAVED')



except Exception as e:
    print(e)  






